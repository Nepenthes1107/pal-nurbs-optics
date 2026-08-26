import shutil
import unittest
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

from biot.domain import Device, ResultStatus, SingleFieldRequest, SystemConfig
from biot.services import compute_single_field, field_angles_to_cb_excel_tilts, modify_excel_config


class TestSingleFieldService(unittest.TestCase):
    def test_field_angles_map_to_cross_axis_coordinate_break_tilts(self):
        self.assertEqual(field_angles_to_cb_excel_tilts(10.0, -25.0), (-25.0, 10.0))

    def test_modify_excel_config_writes_cross_axis_field_tilts(self):
        repo_root = Path(__file__).resolve().parents[1]
        output_dir = repo_root / "results" / "test_field_mapping"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / "mapped_config.xlsx"
        if output_path.exists():
            output_path.unlink()

        modify_excel_config(repo_root / "eye_image_glass.xlsx", output_path, "Infinity", 10.0, -25.0)

        wb = load_workbook(output_path, data_only=True)
        ws = wb.active
        try:
            self.assertEqual(ws["B3"].value, "Infinity")
            self.assertEqual(float(ws["H7"].value), -25.0)
            self.assertEqual(float(ws["I7"].value), 10.0)
        finally:
            wb.close()

    def test_compute_single_field_psf_health(self):
        repo_root = Path(__file__).resolve().parents[1]
        output_dir = repo_root / "results" / "test_service_single_field"
        if output_dir.exists():
            shutil.rmtree(output_dir)
        output_dir.mkdir(parents=True)
        chart_path = output_dir / "chart_input.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        for row in range(1, 9):
            for column in range(1, 9):
                worksheet.cell(row=row, column=column, value=(row + column) % 2)
        workbook.save(chart_path)
        workbook.close()

        req = SingleFieldRequest(
            system=SystemConfig(
                excel_path=repo_root / "eye_image_glass.xlsx",
                object_distance_mm=float("inf"),
                np_pupil=32,
                ni_image=64,
                device=Device.CPU,
            ),
            field_x_deg=0.0,
            field_y_deg=0.0,
            cutoff_cyc_per_mm=100.0,
            with_mtf=True,
            with_chart_convolution=True,
            chart_path=chart_path,
            output_dir=output_dir,
        )

        result = compute_single_field(req)

        self.assertEqual(result.status, ResultStatus.SUCCEEDED, result.error)
        self.assertIsNotNone(result.psf)
        self.assertTrue(np.isfinite(result.psf).all())
        self.assertTrue((result.psf >= 0).all())
        self.assertAlmostEqual(float(result.psf.sum()), 1.0, places=6)
        self.assertAlmostEqual(float(result.metrics["energy_sum"]), 1.0, places=6)
        self.assertIsNotNone(result.mtf_metrics)
        self.assertIsNotNone(result.mtf_curve)
        self.assertEqual(result.mtf_curve.shape[1], 3)
        self.assertIsNotNone(result.chart_image)
        self.assertTrue(np.isfinite(result.chart_image).all())
        self.assertAlmostEqual(float(result.metrics["mtf_dc"]), 1.0, places=12)
        self.assertTrue((output_dir / "psf_data.npy").exists())
        self.assertTrue((output_dir / "psf_data.xlsx").exists())
        self.assertTrue((output_dir / "psf_image.png").exists())
        self.assertTrue((output_dir / "psf_metrics.csv").exists())
        self.assertTrue((output_dir / "mtf_metrics.csv").exists())
        self.assertTrue((output_dir / "chart_convolved.npy").exists())
        self.assertTrue((output_dir / "chart_convolved.png").exists())
        self.assertTrue((output_dir / "manifest.json").exists())


if __name__ == "__main__":
    unittest.main()
