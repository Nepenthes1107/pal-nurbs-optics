from __future__ import annotations

import hashlib
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet import _reader as _openpyxl_ws_reader

_openpyxl_ws_reader._cast_number = lambda value: float(value)

from biot.domain import Device, SystemConfig
from biot.domain.serialization import parse_float


def file_sha256(path: Path) -> str:
    """Return the SHA256 digest for a local file.

    Units: none. This CPU helper reads bytes only and does not participate in
    optical computation or autograd.
    """

    hasher = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def parse_object_distance(value: Any) -> float:
    """Parse an object distance from Excel cell B3.

    Units: returned value is mm; Infinity/inf strings map to float("inf").
    """

    if value is None:
        return float("inf")
    return parse_float(value)


def load_system_from_excel(
    excel_path: Path,
    *,
    device: Device | str = Device.AUTO,
    np_pupil: int = 256,
    ni_image: int = 512,
    zernike_n_max: int = 5,
    wavelength_nm: float = 555.0,
    pupil_radius_mm: float = 2.0,
) -> SystemConfig:
    """Load a minimal `SystemConfig` from the BIOT Excel workbook.

    Excel semantics:
    - B3: object distance in mm or Infinity.
    - H8: lens rotation angle in degree.
    The returned config stores only metadata and does not modify the workbook.
    """

    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 配置文件不存在: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        object_distance_mm = parse_object_distance(ws["B3"].value)
        rotation_raw = ws["H8"].value
        lens_rotation_deg = 0.0 if rotation_raw in (None, "") else float(rotation_raw)
    finally:
        wb.close()

    return SystemConfig(
        excel_path=path,
        object_distance_mm=object_distance_mm,
        excel_sha256=file_sha256(path),
        wavelength_nm=float(wavelength_nm),
        pupil_radius_mm=float(pupil_radius_mm),
        lens_rotation_deg=lens_rotation_deg,
        np_pupil=int(np_pupil),
        ni_image=int(ni_image),
        zernike_n_max=int(zernike_n_max),
        device=Device(device),
    )


def summarize_system(config: SystemConfig) -> dict[str, Any]:
    """Return GUI-friendly system metadata with explicit units."""

    object_distance = "Infinity" if config.object_distance_mm == float("inf") else f"{config.object_distance_mm:g} mm"
    return {
        "excel_path": str(config.excel_path),
        "object_distance": object_distance,
        "wavelength": f"{config.wavelength_nm:g} nm",
        "pupil_radius": f"{config.pupil_radius_mm:g} mm",
        "lens_rotation": f"{config.lens_rotation_deg:g} deg",
        "np_pupil": config.np_pupil,
        "ni_image": config.ni_image,
        "zernike_n_max": config.zernike_n_max,
        "device": config.device.value,
        "excel_sha256": config.excel_sha256,
    }


def validate_system(config: SystemConfig) -> list[str]:
    """Return non-fatal validation issues for a loaded system config."""

    issues: list[str] = []
    if not Path(config.excel_path).exists():
        issues.append(f"Excel 配置文件不存在: {config.excel_path}")
    if config.np_pupil <= 0:
        issues.append("np_pupil 必须为正整数")
    if config.ni_image <= 0:
        issues.append("ni_image 必须为正整数")
    if config.np_pupil > 4096:
        issues.append("np_pupil 超过 4096 上限")
    if config.ni_image > 4096:
        issues.append("ni_image 超过 4096 上限")
    if config.zernike_n_max < 0:
        issues.append("zernike_n_max 必须为非负整数")
    if config.wavelength_nm <= 0:
        issues.append("wavelength_nm 必须为正数")
    if config.pupil_radius_mm <= 0:
        issues.append("pupil_radius_mm 必须为正数")
    return issues


def save_system_config(config: SystemConfig, path: Path) -> Path:
    """Save a `SystemConfig` JSON file."""

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(config.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return target


def load_system_config(path: Path) -> SystemConfig:
    """Load a `SystemConfig` JSON file."""

    data = json.loads(Path(path).read_text(encoding="utf-8"))
    return SystemConfig.from_dict(data)
