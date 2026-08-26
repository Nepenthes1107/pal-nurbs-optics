import os
import subprocess
import sys
import traceback
import tkinter as tk
from pathlib import Path
from tkinter import messagebox

from openpyxl import load_workbook

import conda_env_utils

os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"

PROJECT_ROOT = Path(__file__).resolve().parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

FILE_PATH = PROJECT_ROOT / "eye_image_glass.xlsx"

FIELD_NAMES = [
    "面型",
    "厚度(mm)",
    "半径(mm)",
    "口径(mm)",
    "材料",
    "圆锥系数",
    "4阶项系数",
    "6阶项系数",
    "8阶项系数",
    "10阶项系数",
    "12阶项系数",
    "14阶项系数",
    "16阶项系数",
]

SPECIFIC_FIELD_MAPPING = {
    "旋转角度(度)": (7, 8),
    "玻璃体长度(mm)": (13, 2),
}

ROW_NAMES = {
    5: "镜片1",
    6: "镜片2",
}


def _load_sheet():
    wb = load_workbook(FILE_PATH)
    return wb, wb.active


def update_excel():
    try:
        wb, sheet = _load_sheet()
        selected_row_name = row_var.get()
        selected_row = row_numbers[selected_row_name]

        for col in range(1, 14):
            if selected_row_name == "镜片2" and col == 2:
                continue
            entry = entries[col - 1]
            if entry is None:
                continue
            new_value = entry.get().strip()
            if new_value:
                sheet.cell(row=selected_row, column=col, value=new_value)

        for field, (row, column) in SPECIFIC_FIELD_MAPPING.items():
            new_value = specific_entries[field].get().strip()
            if new_value:
                sheet.cell(row=row, column=column, value=new_value)

        wb.save(FILE_PATH)
        wb.close()
        messagebox.showinfo("成功", "数据已更新")
    except Exception as exc:
        messagebox.showerror("错误", f"数据更新失败: {exc}")


def show_input_fields():
    for widget in input_frame.winfo_children():
        widget.destroy()

    selected_row_name = row_var.get()
    selected_row = row_numbers[selected_row_name]

    wb, sheet = _load_sheet()
    try:
        global entries
        entries = []
        for col in range(1, 14):
            if selected_row_name == "镜片2" and col == 2:
                entries.append(None)
                continue

            col_name = FIELD_NAMES[col - 1]
            current_value = sheet.cell(row=selected_row, column=col).value

            tk.Label(input_frame, text=f"{col_name}:").grid(row=col - 1, column=0, sticky="w")
            entry = tk.Entry(input_frame)
            entry.grid(row=col - 1, column=1)
            if current_value is not None:
                entry.insert(0, str(current_value))
            entries.append(entry)
    finally:
        wb.close()


def run_script_with_cutoff():
    try:
        obj_dist = obj_dist_entry.get().strip()
        field_x = field_x_entry.get().strip()
        field_y = field_y_entry.get().strip()
        cutoff_freq = cutoff_entry.get().strip()
        with_mtf = bool(mtf_var.get())

        if not obj_dist:
            messagebox.showerror("错误", "请输入物距")
            return
        if not field_x:
            messagebox.showerror("错误", "请输入视场角 X")
            return
        if not field_y:
            messagebox.showerror("错误", "请输入视场角 Y")
            return
        if not cutoff_freq:
            messagebox.showerror("错误", "请输入截止频率")
            return

        base_dir = Path(sys._MEIPASS) if getattr(sys, "frozen", False) else PROJECT_ROOT
        excel_path = base_dir / "eye_image_glass.xlsx"

        obj_dist_lower = obj_dist.lower()
        obj_dist_str = obj_dist if obj_dist_lower in {"inf", "infinity"} else f"{obj_dist}mm"
        output_dir = f"gui_results/{obj_dist_str}_field({field_x},{field_y})"

        if getattr(sys, "frozen", False):
            cmd = [
                str(base_dir / "multi_rays.exe"),
                str(excel_path),
                obj_dist,
                field_x,
                field_y,
                "--cutoff",
                cutoff_freq,
                "--output",
                output_dir,
            ]
        else:
            cmd = [
                conda_env_utils.python_exe(),
                str(base_dir / "multi_rays.py"),
                str(excel_path),
                obj_dist,
                field_x,
                field_y,
                "--cutoff",
                cutoff_freq,
                "--output",
                output_dir,
            ]

        if with_mtf:
            cmd.append("--with-mtf")

        conda_env_utils.run(
            cmd,
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )

        mode_text = "PSF + MTF" if with_mtf else "仅 PSF"
        messagebox.showinfo(
            "计算成功",
            f"{mode_text} 计算完成\n\n"
            f"配置:\n"
            f"  物距: {obj_dist} mm\n"
            f"  视场: ({field_x}, {field_y}) 度\n"
            f"  截止频率: {cutoff_freq} cycles/mm\n\n"
            f"结果已保存到: {output_dir}",
        )
    except subprocess.CalledProcessError as exc:
        error_msg = (
            f"计算失败 (返回码 {exc.returncode})\n\n"
            f"命令: {' '.join(exc.cmd)}\n\n"
            f"错误输出:\n{exc.stderr}\n\n"
            f"标准输出:\n{exc.stdout}"
        )
        messagebox.showerror("计算错误", error_msg)
    except Exception as exc:
        error_msg = f"意外错误:\n{exc}\n\n堆栈信息:\n{traceback.format_exc()}"
        messagebox.showerror("系统错误", error_msg)


root = tk.Tk()
root.title("计算 FFT PSF")
root.geometry("900x700")
root.minsize(700, 500)
root.maxsize(1200, 1000)
root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

row_numbers = {name: row for row, name in ROW_NAMES.items()}

row_var = tk.StringVar(value="镜片1")
row_menu = tk.OptionMenu(root, row_var, *ROW_NAMES.values(), command=lambda _: show_input_fields())
row_menu.grid(row=0, column=0, columnspan=2, pady=10)

input_frame = tk.Frame(root)
input_frame.grid(row=1, column=0, columnspan=2)

specific_frame = tk.Frame(root)
specific_frame.grid(row=2, column=0, columnspan=2, pady=10)

specific_entries = {}
for i, (field, (row, column)) in enumerate(SPECIFIC_FIELD_MAPPING.items()):
    tk.Label(specific_frame, text=f"{field}:").grid(row=i, column=0, sticky="w")
    entry = tk.Entry(specific_frame)
    entry.grid(row=i, column=1)
    specific_entries[field] = entry

params_frame = tk.Frame(root)
params_frame.grid(row=3, column=0, columnspan=2, pady=10)

tk.Label(params_frame, text="物距 (mm，输入 inf 表示无穷远):").grid(row=0, column=0, sticky="w", padx=5)
obj_dist_entry = tk.Entry(params_frame, width=15)
obj_dist_entry.insert(0, "inf")
obj_dist_entry.grid(row=0, column=1, padx=5)

tk.Label(params_frame, text="视场角 X (度):").grid(row=1, column=0, sticky="w", padx=5)
field_x_entry = tk.Entry(params_frame, width=15)
field_x_entry.insert(0, "0")
field_x_entry.grid(row=1, column=1, padx=5)

tk.Label(params_frame, text="视场角 Y (度):").grid(row=2, column=0, sticky="w", padx=5)
field_y_entry = tk.Entry(params_frame, width=15)
field_y_entry.insert(0, "0")
field_y_entry.grid(row=2, column=1, padx=5)

tk.Label(params_frame, text="MTF 截止频率 (cycles/mm):").grid(row=3, column=0, sticky="w", padx=5)
cutoff_entry = tk.Entry(params_frame, width=15)
cutoff_entry.insert(0, "100")
cutoff_entry.grid(row=3, column=1, padx=5)

update_button = tk.Button(root, text="更新镜片参数", command=update_excel)
update_button.grid(row=4, column=0, columnspan=2, pady=10)

run_with_cutoff_button = tk.Button(root, text="计算 FFT PSF", command=run_script_with_cutoff)
run_with_cutoff_button.grid(row=5, column=0, pady=10, sticky="e", padx=6)

mtf_var = tk.BooleanVar(value=False)
mtf_checkbox = tk.Checkbutton(root, text="MTF", variable=mtf_var)
mtf_checkbox.grid(row=5, column=1, pady=10, sticky="w", padx=6)

entries = []
show_input_fields()

root.mainloop()
